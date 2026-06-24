"""
Dashboard Corporativo de Vendas — Coco & Cia (v3)
=================================================
Mudanças desta versão (pós-reunião 2):
  + Aba "Diagnóstico Executivo" (relatório em uma página)
  + Aba "Comparação" (lado-a-lado entre clientes ou estados)
  + Aba "Preço Médio" (detecção de distorção entre clientes)
  + R$/kg por linha (extrai peso da descrição; alerta contra-senso)
  + Cliente e produto agora exibidos como "Código - Nome"
  + Curva ABC mantida e em destaque no diagnóstico
  + Exportação PDF começa com a página de Diagnóstico Executivo

Execução:
    streamlit run dashboard_vendas.py
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage,
)

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
st.set_page_config(
    page_title="Dashboard Comercial — Coco & Cia",
    page_icon="📊", layout="wide", initial_sidebar_state="expanded",
)
st.markdown(
    """
    <style>
        .main .block-container {padding-top: 1.2rem; padding-bottom: 2rem;}
        [data-testid="stMetricValue"] {font-size: 1.4rem;}
        [data-testid="stMetricLabel"] {font-weight: 600; font-size: 0.85rem;}
        h1, h2, h3 {color: #102A43;}
        .stTabs [data-baseweb="tab-list"] {gap: 4px; flex-wrap: wrap;}
        .stTabs [data-baseweb="tab"] {padding: 8px 14px; border-radius: 6px 6px 0 0;
                                       background-color: #F0F4F8; font-size: 0.88rem;}
        .stTabs [aria-selected="true"] {background-color: #0B5FFF; color: white;}
        div[data-testid="stSidebarUserContent"] {padding-top: 1rem;}
        .alert-box {padding: 10px 14px; border-radius: 6px; margin: 6px 0;
                    border-left: 4px solid; font-size: 0.9rem;}
        .alert-danger {background:#FEF2F2; border-color:#D64545; color:#7B1818;}
        .alert-warn   {background:#FFF8E1; border-color:#F0B429; color:#8A5800;}
        .alert-ok     {background:#E6F6F0; border-color:#27AB83; color:#0E5E3F;}
    </style>
    """,
    unsafe_allow_html=True,
)

COLORS = {
    "Venda": "#0B5FFF", "Devolução": "#D64545", "Líquido": "#27AB83",
    "Úmido": "#2680C2", "Integral": "#F0B429",
    "Varejo": "#0B5FFF", "Food": "#F0B429", "Profissional": "#7B61FF",
    "Outros": "#9AA5B1",
    "A": "#27AB83", "B": "#F0B429", "C": "#D64545",
    "2025": "#9AA5B1", "2026": "#0B5FFF",
}
MESES_PT = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
            7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

# Paleta para múltiplas entidades na aba de Comparação
PALETA_COMPARE = ["#0B5FFF", "#F0B429", "#27AB83", "#D64545", "#7B61FF", "#2680C2"]


# =============================================================================
# FORMATAÇÃO
# =============================================================================
def fmt_brl(v):
    if pd.isna(v): return "R$ 0,00"
    return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

def fmt_int(v):
    if pd.isna(v): return "0"
    return f"{v:,.0f}".replace(",",".")

def fmt_pct(v):
    if pd.isna(v): return "0,0%"
    return f"{v*100:,.1f}%".replace(".",",")

def fmt_delta_pct(v):
    if pd.isna(v): return "—"
    sinal = "+" if v >= 0 else ""
    return f"{sinal}{v*100:,.1f}%".replace(".",",")

def fmt_rkg(v):
    if pd.isna(v) or v == 0: return "—"
    return f"R$ {v:,.2f}/kg".replace(",","X").replace(".",",").replace("X",".")


# =============================================================================
# CLASSIFICAÇÃO DE PRODUTOS + EXTRAÇÃO DE PESO
# =============================================================================
def _norm(t):
    if not isinstance(t, str): return ""
    return "".join(c for c in unicodedata.normalize("NFKD", t) if not unicodedata.combining(c)).upper()

def classify_linha(d):
    d = _norm(d)
    if not d: return "Outros"
    if re.search(r"\b5\s*KG\b", d): return "Profissional"
    if re.search(r"10\s*X\s*1\s*KG", d) or re.search(r"\b3\s*L\b", d): return "Food"
    return "Varejo"

def classify_categoria(d):
    d = _norm(d)
    if not d: return "Outros"
    is_ral = "RAL" in d; is_floc = "FLOC" in d
    is_umi = "UMI" in d
    is_int = "INT" in d or "INTEGRAL" in d or "QUEIMADO" in d
    if is_ral and is_umi: return "Ralado úmido"
    if is_ral and is_int: return "Ralado integral"
    if is_floc and is_umi: return "Flocado úmido"
    if is_floc and is_int: return "Flocado integral"
    if "LEITE" in d and "ESPECIAL" in d: return "Leite especial"
    if "LEITE" in d and "PROFISSIONAL" in d: return "Leite profissional"
    if "OLEO" in d: return "Óleo"
    return "Outros"

def classify_materia_prima(d):
    d = _norm(d)
    if not d: return "N/A"
    if "UMI" in d: return "Úmido"
    if "INT" in d or "INTEGRAL" in d or "QUEIMADO" in d: return "Integral"
    return "N/A"

def extract_peso_kg(descricao):
    """
    Extrai peso total da embalagem em kg, a partir da descrição.
    Trata líquidos (L/ML) como kg (densidade ≈1, aproximação aceitável p/ leite/óleo).
    Padrões suportados: 24X100G, 10X1KG, 5KG, 24X200ML, 3L, etc.
    """
    if not isinstance(descricao, str): return None
    d = _norm(descricao)
    m = re.search(r"(\d+)\s*X\s*(\d+(?:[.,]\d+)?)\s*(KG|G|ML|L)\b", d)
    if m:
        n = int(m.group(1)); val = float(m.group(2).replace(",", "."))
        unit = m.group(3)
        if unit == "KG": return n * val
        if unit == "G":  return n * val / 1000
        if unit == "L":  return n * val
        if unit == "ML": return n * val / 1000
    m = re.search(r"\b(\d+(?:[.,]\d+)?)\s*(KG|L)\b", d)
    if m:
        return float(m.group(1).replace(",", "."))
    return None


# =============================================================================
# LOADERS
# =============================================================================
@st.cache_data(show_spinner=False)
def load_vendas(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=1)
    req = ["Cliente", "Produto", "Emissao", "Quantidade", "Vlr.Total"]
    miss = [c for c in req if c not in df.columns]
    if miss: raise ValueError(f"Colunas ausentes em VENDAS: {miss}")
    df = df.dropna(subset=["Emissao","Cliente","Produto"]).copy()
    df["Emissao"] = pd.to_datetime(df["Emissao"], errors="coerce")
    df = df.dropna(subset=["Emissao"])
    df["Cliente"] = df["Cliente"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["Produto"] = df["Produto"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["Descricao"] = df.get("Descricao","").astype(str).str.strip()
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce").fillna(0.0)
    df["Vlr.Total"] = pd.to_numeric(df["Vlr.Total"], errors="coerce").fillna(0.0)
    if "Vendedor" in df.columns:
        df["Vendedor"] = pd.to_numeric(df["Vendedor"], errors="coerce").fillna(0).astype(int).astype(str)
        df["Vendedor"] = df["Vendedor"].replace("0","(Sem vendedor)")
    else:
        df["Vendedor"] = "(Sem vendedor)"
    df["AnoMes"] = df["Emissao"].dt.to_period("M").dt.to_timestamp()
    df["Ano"] = df["Emissao"].dt.year.astype(int)
    df["Mes"] = df["Emissao"].dt.month.astype(int)
    df["Tipo"] = "Venda"
    return df

@st.cache_data(show_spinner=False)
def load_devolucoes(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=1)
    req = ["Forn/Cliente","Produto","DT Digitacao","Quantidade","Vlr.Total"]
    miss = [c for c in req if c not in df.columns]
    if miss: raise ValueError(f"Colunas ausentes em DEVOLUÇÕES: {miss}")
    df = df.dropna(subset=["DT Digitacao","Forn/Cliente","Produto"]).copy()
    df = df.rename(columns={"Forn/Cliente":"Cliente","DT Digitacao":"Emissao"})
    df["Emissao"] = pd.to_datetime(df["Emissao"], errors="coerce")
    df = df.dropna(subset=["Emissao"])
    df["Cliente"] = df["Cliente"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["Produto"] = df["Produto"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["Descricao"] = df.get("Descricao","").astype(str).str.strip()
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce").fillna(0.0)
    df["Vlr.Total"] = pd.to_numeric(df["Vlr.Total"], errors="coerce").fillna(0.0)
    df["Vendedor"] = "(N/A)"
    # Regra contábil crítica: AnoMes vem da DT Digitacao
    df["AnoMes"] = df["Emissao"].dt.to_period("M").dt.to_timestamp()
    df["Ano"] = df["Emissao"].dt.year.astype(int)
    df["Mes"] = df["Emissao"].dt.month.astype(int)
    df["Tipo"] = "Devolução"
    return df

@st.cache_data(show_spinner=False)
def load_clientes(file_bytes):
    raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    header_row = None
    for i in range(min(10, len(raw))):
        row = raw.iloc[i].astype(str).str.strip().tolist()
        if "Codigo" in row and "Nome" in row:
            header_row = i; break
    if header_row is None:
        raise ValueError("Cabeçalho do BANCO_DE_CLIENTES não localizado.")
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=header_row)
    keep = [c for c in ["Codigo","Nome","Estado","Municipio","Regiao","Desc.Região"] if c in df.columns]
    df = df[keep].copy()
    df = df.dropna(subset=["Codigo"])
    df["Codigo"] = df["Codigo"].astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df = df[df["Codigo"].str.match(r"^\d+$", na=False)]
    for c in ["Nome","Estado","Municipio","Desc.Região"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().replace({"nan": None})
    df["Regiao_Desc"] = df.get("Desc.Região", df.get("Regiao", pd.Series(dtype=str))).fillna("(Sem região)")
    df["Estado"] = df.get("Estado", pd.Series(dtype=str)).fillna("(Sem UF)")
    df["Municipio"] = df.get("Municipio", pd.Series(dtype=str)).fillna("(Sem município)")
    df["Nome"] = df.get("Nome", pd.Series(dtype=str)).fillna("(Sem nome)")
    return df.drop_duplicates(subset="Codigo", keep="first")

def enrich(df, clientes):
    out = df.merge(
        clientes[["Codigo","Nome","Estado","Municipio","Regiao_Desc"]],
        left_on="Cliente", right_on="Codigo", how="left", suffixes=("","_cli"))
    out["Nome"] = out["Nome"].fillna("(Cliente sem cadastro)")
    out["Estado"] = out["Estado"].fillna("(Sem UF)")
    out["Municipio"] = out["Municipio"].fillna("(Sem município)")
    out["Regiao_Desc"] = out["Regiao_Desc"].fillna("(Sem região)")

    # Pedido da reunião: cliente e produto exibidos como "Código - Nome/Desc"
    out["Cliente_Label"] = out["Cliente"].astype(str) + " — " + out["Nome"].astype(str)
    out["Produto_Label"] = out["Produto"].astype(str) + " — " + out["Descricao"].astype(str)

    out["Linha"] = out["Descricao"].apply(classify_linha)
    out["Categoria"] = out["Descricao"].apply(classify_categoria)
    out["Materia_Prima"] = out["Descricao"].apply(classify_materia_prima)
    out["Peso_kg_pack"] = out["Descricao"].apply(extract_peso_kg)
    # Peso total da linha vendida = peso_pack × quantidade (em caixas/fardos)
    out["Peso_kg_total"] = out["Peso_kg_pack"] * out["Quantidade"]
    return out


# =============================================================================
# ANÁLISES (funções puras, reutilizadas pela UI e pela exportação)
# =============================================================================
def calc_kpis(vendas_f, devol_f):
    v = vendas_f["Vlr.Total"].sum(); d = devol_f["Vlr.Total"].sum()
    qv = vendas_f["Quantidade"].sum(); qd = devol_f["Quantidade"].sum()
    peso = vendas_f["Peso_kg_total"].sum()
    return {
        "bruto": v, "devol_vlr": d, "liquido": v-d,
        "qtd_v": qv, "qtd_d": qd, "qtd_liq": qv-qd,
        "pct_devol": (d/v) if v else 0,
        "clientes_ativos": vendas_f["Cliente"].nunique(),
        "produtos_vendidos": vendas_f["Produto"].nunique(),
        "preco_medio_unit": (v/qv) if qv else 0,
        "preco_medio_kg": (v/peso) if peso else 0,
        "peso_kg": peso,
    }

def analise_abc(vendas_f, devol_f):
    v = vendas_f.groupby(["Produto","Descricao"], as_index=False)["Vlr.Total"].sum()
    d = devol_f.groupby("Produto", as_index=False)["Vlr.Total"].sum()
    abc = v.merge(d, on="Produto", how="left", suffixes=("_v","_d")).fillna(0)
    abc["Líquido"] = abc["Vlr.Total_v"] - abc["Vlr.Total_d"]
    abc = abc.sort_values("Líquido", ascending=False).reset_index(drop=True)
    total = abc["Líquido"].sum()
    if total <= 0: return abc, total
    abc["%"] = abc["Líquido"] / total
    abc["% Acumulado"] = abc["%"].cumsum()
    abc["Classe"] = abc["% Acumulado"].apply(lambda p: "A" if p<=0.80 else ("B" if p<=0.95 else "C"))
    abc["Produto_Label"] = abc["Produto"].astype(str) + " — " + abc["Descricao"].astype(str)
    return abc, total

def clientes_em_risco(vendas_f, anos_target=(2025, 2026)):
    if not all(a in vendas_f["Ano"].unique() for a in anos_target):
        return pd.DataFrame()
    a1, a2 = anos_target
    meses_a2 = sorted(vendas_f[vendas_f["Ano"]==a2]["Mes"].unique())
    if not meses_a2: return pd.DataFrame()
    base = vendas_f[vendas_f["Mes"].isin(meses_a2)]
    piv = base.pivot_table(
        index=["Cliente","Nome","Regiao_Desc","Estado"],
        columns="Ano", values="Vlr.Total", aggfunc="sum", fill_value=0,
    ).reset_index()
    if a1 not in piv.columns or a2 not in piv.columns:
        return pd.DataFrame()
    piv = piv[piv[a1] >= 1000].copy()
    piv["Δ Abs"] = piv[a2] - piv[a1]
    piv["Δ %"] = (piv[a2] - piv[a1]) / piv[a1].replace(0, pd.NA)
    risco = piv[piv["Δ %"].fillna(-1) <= -0.30].copy()
    def nivel(r):
        if r[a2] == 0: return "🔴 Crítico (parou)"
        if r["Δ %"] <= -0.70: return "🔴 Severo"
        if r["Δ %"] <= -0.50: return "🟠 Alto"
        return "🟡 Moderado"
    if not risco.empty:
        risco["Nível"] = risco.apply(nivel, axis=1)
        risco["Cliente_Label"] = risco["Cliente"].astype(str) + " — " + risco["Nome"].astype(str)
        risco = risco.sort_values("Δ Abs", ascending=True)
    return risco

def crescimento_geo(vendas_f, devol_f, dim="Regiao_Desc", anos_target=(2025, 2026)):
    a1, a2 = anos_target
    if not all(a in vendas_f["Ano"].unique() for a in anos_target):
        return pd.DataFrame()
    comuns = sorted(set(vendas_f[vendas_f["Ano"]==a1]["Mes"]) & set(vendas_f[vendas_f["Ano"]==a2]["Mes"]))
    if not comuns: return pd.DataFrame()
    v = vendas_f[vendas_f["Mes"].isin(comuns)]
    d = devol_f[devol_f["Mes"].isin(comuns)]
    gv = v.groupby([dim,"Ano"])["Vlr.Total"].sum().unstack(fill_value=0)
    gd = d.groupby([dim,"Ano"])["Vlr.Total"].sum().unstack(fill_value=0)
    for a in anos_target:
        if a not in gv.columns: gv[a] = 0
        if a not in gd.columns: gd[a] = 0
    liq = gv - gd
    liq["Δ Abs"] = liq[a2] - liq[a1]
    liq["Δ %"] = (liq[a2] - liq[a1]) / liq[a1].replace(0, pd.NA)
    return liq.reset_index().sort_values("Δ Abs", ascending=False)

def devolucoes_problematicas(vendas_f, devol_f):
    vp = vendas_f.groupby(["Produto","Descricao"], as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Vendido"})
    dp = devol_f.groupby("Produto", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Devolvido"})
    prod = vp.merge(dp, on="Produto", how="left").fillna(0)
    prod = prod[prod["Vendido"] >= 5000].copy()
    prod["Taxa"] = prod["Devolvido"] / prod["Vendido"]
    prod["Produto_Label"] = prod["Produto"].astype(str) + " — " + prod["Descricao"].astype(str)
    prod = prod.sort_values("Taxa", ascending=False)

    vc = vendas_f.groupby(["Cliente","Nome"], as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Vendido"})
    dc = devol_f.groupby("Cliente", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Devolvido"})
    cli = vc.merge(dc, on="Cliente", how="left").fillna(0)
    cli = cli[cli["Vendido"] >= 5000].copy()
    cli["Taxa"] = cli["Devolvido"] / cli["Vendido"]
    cli["Cliente_Label"] = cli["Cliente"].astype(str) + " — " + cli["Nome"].astype(str)
    cli = cli.sort_values("Taxa", ascending=False)
    return prod, cli

def ranking_vendedores(vendas_f):
    if "Vendedor" not in vendas_f.columns: return pd.DataFrame()
    rk = vendas_f.groupby("Vendedor", as_index=False).agg(
        Faturamento=("Vlr.Total","sum"),
        Pedidos=("Num. Docto.","nunique") if "Num. Docto." in vendas_f.columns else ("Vlr.Total","count"),
        Quantidade=("Quantidade","sum"))
    rk["Ticket_Medio"] = rk["Faturamento"] / rk["Pedidos"].replace(0, pd.NA)
    return rk.sort_values("Faturamento", ascending=False)

def sazonalidade(vendas_f, devol_f):
    v = vendas_f.groupby(["Ano","Mes"])["Vlr.Total"].sum().reset_index()
    d = devol_f.groupby(["Ano","Mes"])["Vlr.Total"].sum().reset_index()
    m = v.merge(d, on=["Ano","Mes"], how="outer", suffixes=("_v","_d")).fillna(0)
    m["Líquido"] = m["Vlr.Total_v"] - m["Vlr.Total_d"]
    pivot = m.pivot(index="Ano", columns="Mes", values="Líquido").fillna(0)
    pivot = pivot.reindex(columns=range(1,13), fill_value=0)
    media = pivot.replace(0, pd.NA).mean(axis=0).reset_index()
    media.columns = ["Mes","Media_Liquido"]
    media["MesNome"] = media["Mes"].map(MESES_PT)
    return pivot, media

# ---- NOVAS análises pedidas na reunião 2 ----

def preco_medio_por_produto(vendas_f):
    """Preço médio unitário (R$/CX) por produto + dispersão entre clientes."""
    if vendas_f.empty: return pd.DataFrame()
    g = vendas_f.groupby(["Produto","Descricao"]).agg(
        Faturamento=("Vlr.Total","sum"),
        Quantidade=("Quantidade","sum")).reset_index()
    g["Preço_Medio_CX"] = g["Faturamento"] / g["Quantidade"].replace(0, pd.NA)
    g["Produto_Label"] = g["Produto"].astype(str) + " — " + g["Descricao"].astype(str)

    # Dispersão de preço por cliente (para o MESMO produto)
    pc = vendas_f.groupby(["Produto","Cliente","Nome"]).agg(
        F=("Vlr.Total","sum"), Q=("Quantidade","sum")).reset_index()
    pc["Preco_CX"] = pc["F"] / pc["Q"].replace(0, pd.NA)
    stats = pc.groupby("Produto")["Preco_CX"].agg(
        Min="min", Max="max", Std="std", N="count").reset_index()
    stats["Amplitude_%"] = (stats["Max"]-stats["Min"]) / stats["Min"].replace(0, pd.NA)
    out = g.merge(stats, on="Produto", how="left")
    return out.sort_values("Faturamento", ascending=False)

def distorcao_preco_cliente(vendas_f, produto_codigo):
    """Para um produto específico, mostra preço médio que cada cliente pagou."""
    base = vendas_f[vendas_f["Produto"] == produto_codigo]
    if base.empty: return pd.DataFrame()
    g = base.groupby(["Cliente","Nome","Regiao_Desc","Estado"]).agg(
        Faturamento=("Vlr.Total","sum"),
        Quantidade=("Quantidade","sum")).reset_index()
    g["Preco_CX"] = g["Faturamento"] / g["Quantidade"].replace(0, pd.NA)
    g["Cliente_Label"] = g["Cliente"].astype(str) + " — " + g["Nome"].astype(str)
    g["% vs Mediana"] = (g["Preco_CX"] / g["Preco_CX"].median() - 1)
    return g.sort_values("Preco_CX", ascending=False)

def preco_kg_por_linha(vendas_f):
    """R$/kg médio por Linha e Categoria. Alerta se Food < Profissional."""
    base = vendas_f[vendas_f["Peso_kg_pack"].notna() & (vendas_f["Peso_kg_pack"] > 0)]
    if base.empty: return pd.DataFrame(), pd.DataFrame(), False, ""
    # Por linha
    por_linha = base.groupby("Linha").agg(
        Faturamento=("Vlr.Total","sum"),
        Peso_kg=("Peso_kg_total","sum"),
        Quantidade=("Quantidade","sum")).reset_index()
    por_linha["R$/kg"] = por_linha["Faturamento"] / por_linha["Peso_kg"].replace(0, pd.NA)
    # Por categoria
    por_cat = base.groupby(["Linha","Categoria"]).agg(
        Faturamento=("Vlr.Total","sum"),
        Peso_kg=("Peso_kg_total","sum")).reset_index()
    por_cat["R$/kg"] = por_cat["Faturamento"] / por_cat["Peso_kg"].replace(0, pd.NA)

    # Detecção do contra-senso: Food < Profissional (deveria ser o oposto)
    alerta = False; texto = ""
    food = por_linha[por_linha["Linha"]=="Food"]["R$/kg"]
    prof = por_linha[por_linha["Linha"]=="Profissional"]["R$/kg"]
    if not food.empty and not prof.empty:
        rf, rp = food.iloc[0], prof.iloc[0]
        if pd.notna(rf) and pd.notna(rp) and rf < rp:
            alerta = True
            texto = (f"⚠️ Contra-senso detectado: Food sai a {fmt_rkg(rf)} "
                     f"e Profissional a {fmt_rkg(rp)}. Em tese, embalagem maior "
                     "(Profissional 5kg) deveria custar menos por kg que Food (10×1kg).")
    return por_linha, por_cat, alerta, texto

def comparar_entidades(vendas_f, devol_f, dim, entidades):
    """Calcula resumo lado-a-lado de N entidades (clientes ou estados)."""
    if not entidades: return None
    res = []
    for ent in entidades:
        v = vendas_f[vendas_f[dim] == ent]
        d = devol_f[devol_f[dim] == ent]
        kpi = calc_kpis(v, d)
        res.append({
            "Entidade": ent,
            "Faturamento Bruto": kpi["bruto"],
            "Devoluções": kpi["devol_vlr"],
            "Líquido": kpi["liquido"],
            "Qtd": kpi["qtd_v"],
            "% Devol.": kpi["pct_devol"],
            "Preço Médio (R$/CX)": kpi["preco_medio_unit"],
            "Preço Médio (R$/kg)": kpi["preco_medio_kg"],
            "Clientes Ativos": kpi["clientes_ativos"],
            "SKUs": kpi["produtos_vendidos"],
        })
    return pd.DataFrame(res)


# =============================================================================
# EXPORTAÇÃO EXCEL
# =============================================================================
def gerar_excel(vendas_f, devol_f, filtros_desc):
    out = io.BytesIO()
    wb = Workbook()
    azul = PatternFill("solid", fgColor="0B5FFF")
    branco = Font(bold=True, color="FFFFFF", size=11)

    def estilo(ws, n):
        for c in range(1, n+1):
            cell = ws.cell(row=1, column=c)
            cell.fill = azul; cell.font = branco
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(c)].width = 22

    def add_aba(nome, df, fmt=None):
        ws = wb.create_sheet(nome[:31])
        if df.empty:
            ws["A1"] = "(sem dados)"; return
        for i, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=i, value=str(col))
        for r, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                if fmt and df.columns[ci-1] in fmt:
                    val = fmt[df.columns[ci-1]](val)
                ws.cell(row=r, column=ci, value=val)
        estilo(ws, len(df.columns))

    # Diagnóstico
    ws = wb.active; ws.title = "Diagnóstico"
    ws["A1"] = "Diagnóstico Executivo — Coco & Cia"
    ws["A1"].font = Font(bold=True, size=16, color="102A43")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"] = f"Filtros: {filtros_desc or 'Nenhum'}"
    kpi = calc_kpis(vendas_f, devol_f)
    ws["A5"] = "Indicador"; ws["B5"] = "Valor"
    ws["A5"].font = branco; ws["B5"].font = branco
    ws["A5"].fill = azul; ws["B5"].fill = azul
    linhas_kpi = [
        ("Faturamento Bruto", fmt_brl(kpi["bruto"])),
        ("Devoluções (R$)", fmt_brl(kpi["devol_vlr"])),
        ("Faturamento Líquido", fmt_brl(kpi["liquido"])),
        ("% Devolução", fmt_pct(kpi["pct_devol"])),
        ("Quantidade Vendida (CX)", fmt_int(kpi["qtd_v"])),
        ("Quantidade Devolvida (CX)", fmt_int(kpi["qtd_d"])),
        ("Peso Total Vendido (kg)", fmt_int(kpi["peso_kg"])),
        ("Preço Médio por CX", fmt_brl(kpi["preco_medio_unit"])),
        ("Preço Médio por kg", fmt_rkg(kpi["preco_medio_kg"])),
        ("Clientes Ativos", fmt_int(kpi["clientes_ativos"])),
        ("SKUs Vendidos", fmt_int(kpi["produtos_vendidos"])),
    ]
    for i, (k, v) in enumerate(linhas_kpi, 6):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 24

    # Mensal
    mv = vendas_f.groupby("AnoMes", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Vendas"})
    md = devol_f.groupby("AnoMes", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"Devolucoes"})
    mensal = mv.merge(md, on="AnoMes", how="outer").fillna(0)
    mensal["Liquido"] = mensal["Vendas"] - mensal["Devolucoes"]
    mensal["AnoMes"] = mensal["AnoMes"].dt.strftime("%b/%Y")
    add_aba("Mensal", mensal, {"Vendas":fmt_brl,"Devolucoes":fmt_brl,"Liquido":fmt_brl})

    # Categorias e Linhas
    cat = vendas_f.groupby("Categoria", as_index=False).agg(
        Faturamento=("Vlr.Total","sum"), Quantidade=("Quantidade","sum")).sort_values("Faturamento", ascending=False)
    add_aba("Por Categoria", cat, {"Faturamento":fmt_brl,"Quantidade":fmt_int})

    linha_d = vendas_f.groupby("Linha", as_index=False).agg(
        Faturamento=("Vlr.Total","sum"), Quantidade=("Quantidade","sum"),
        Peso_kg=("Peso_kg_total","sum"))
    linha_d["R$/kg"] = linha_d["Faturamento"] / linha_d["Peso_kg"].replace(0, pd.NA)
    add_aba("Por Linha", linha_d, {"Faturamento":fmt_brl,"Quantidade":fmt_int,"Peso_kg":fmt_int,"R$/kg":fmt_rkg})

    # ABC
    abc, _ = analise_abc(vendas_f, devol_f)
    if not abc.empty:
        out_abc = abc[["Produto_Label","Vlr.Total_v","Vlr.Total_d","Líquido","%","% Acumulado","Classe"]]
        out_abc.columns = ["Produto","Bruto","Devoluções","Líquido","%","% Acum.","Classe"]
        add_aba("Curva ABC", out_abc, {
            "Bruto":fmt_brl,"Devoluções":fmt_brl,"Líquido":fmt_brl,
            "%":fmt_pct,"% Acum.":fmt_pct})

    # Top 30 Clientes
    tcv = vendas_f.groupby(["Cliente","Nome","Regiao_Desc","Estado"], as_index=False)["Vlr.Total"].sum()
    tcd = devol_f.groupby("Cliente", as_index=False)["Vlr.Total"].sum()
    tc = tcv.merge(tcd, on="Cliente", how="left", suffixes=("_v","_d")).fillna(0)
    tc["Líquido"] = tc["Vlr.Total_v"] - tc["Vlr.Total_d"]
    tc = tc.nlargest(30, "Líquido")
    tc["Cliente"] = tc["Cliente"].astype(str) + " — " + tc["Nome"].astype(str)
    tc_out = tc[["Cliente","Regiao_Desc","Estado","Vlr.Total_v","Vlr.Total_d","Líquido"]]
    tc_out.columns = ["Cliente","Região","UF","Bruto","Devoluções","Líquido"]
    add_aba("Top 30 Clientes", tc_out, {"Bruto":fmt_brl,"Devoluções":fmt_brl,"Líquido":fmt_brl})

    # YoY
    if {2025, 2026}.issubset(set(vendas_f["Ano"].unique())):
        comuns = sorted(set(vendas_f[vendas_f["Ano"]==2025]["Mes"]) & set(vendas_f[vendas_f["Ano"]==2026]["Mes"]))
        v25 = vendas_f[(vendas_f["Ano"]==2025)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        v26 = vendas_f[(vendas_f["Ano"]==2026)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        d25 = devol_f[(devol_f["Ano"]==2025)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        d26 = devol_f[(devol_f["Ano"]==2026)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        yoy_df = pd.DataFrame({
            "Indicador":["Bruto","Devoluções","Líquido"],
            "2025":[v25, d25, v25-d25], "2026":[v26, d26, v26-d26]})
        yoy_df["Δ Abs"] = yoy_df["2026"] - yoy_df["2025"]
        yoy_df["Δ %"] = (yoy_df["2026"]-yoy_df["2025"]) / yoy_df["2025"].replace(0, pd.NA)
        add_aba("YoY 2025 vs 2026", yoy_df, {
            "2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct})

    # Clientes em risco
    risco = clientes_em_risco(vendas_f)
    if not risco.empty:
        r2 = risco.copy()
        if 2025 in r2.columns: r2 = r2.rename(columns={2025:"Fat 2025"})
        if 2026 in r2.columns: r2 = r2.rename(columns={2026:"Fat 2026"})
        cols = ["Cliente_Label","Regiao_Desc","Estado","Fat 2025","Fat 2026","Δ Abs","Δ %","Nível"]
        cols = [c for c in cols if c in r2.columns]
        r2 = r2[cols].rename(columns={"Cliente_Label":"Cliente"})
        add_aba("Clientes em Risco", r2, {
            "Fat 2025":fmt_brl,"Fat 2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct})

    # Crescimento
    gr = crescimento_geo(vendas_f, devol_f, "Regiao_Desc")
    if not gr.empty:
        gr.columns = [str(c) for c in gr.columns]
        add_aba("Crescimento Região", gr, {"2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct})
    gu = crescimento_geo(vendas_f, devol_f, "Estado")
    if not gu.empty:
        gu.columns = [str(c) for c in gu.columns]
        add_aba("Crescimento UF", gu, {"2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct})

    # Devoluções problemáticas
    pp, pc = devolucoes_problematicas(vendas_f, devol_f)
    if not pp.empty:
        pp_out = pp[["Produto_Label","Vendido","Devolvido","Taxa"]].head(30).rename(columns={"Produto_Label":"Produto"})
        add_aba("Devoluções Produtos", pp_out, {"Vendido":fmt_brl,"Devolvido":fmt_brl,"Taxa":fmt_pct})
    if not pc.empty:
        pc_out = pc[["Cliente_Label","Vendido","Devolvido","Taxa"]].head(30).rename(columns={"Cliente_Label":"Cliente"})
        add_aba("Devoluções Clientes", pc_out, {"Vendido":fmt_brl,"Devolvido":fmt_brl,"Taxa":fmt_pct})

    # Vendedores
    vend = ranking_vendedores(vendas_f)
    if not vend.empty:
        add_aba("Vendedores", vend, {
            "Faturamento":fmt_brl,"Quantidade":fmt_int,"Pedidos":fmt_int,"Ticket_Medio":fmt_brl})

    # Preço Médio por Produto
    pm = preco_medio_por_produto(vendas_f)
    if not pm.empty:
        pm_out = pm[["Produto_Label","Faturamento","Quantidade","Preço_Medio_CX","Min","Max","Amplitude_%","N"]]
        pm_out.columns = ["Produto","Faturamento","Qtd","Preço Médio (CX)","Mín. cliente","Máx. cliente","Amplitude %","Nº clientes"]
        add_aba("Preço Médio Produto", pm_out, {
            "Faturamento":fmt_brl,"Qtd":fmt_int,"Preço Médio (CX)":fmt_brl,
            "Mín. cliente":fmt_brl,"Máx. cliente":fmt_brl,"Amplitude %":fmt_pct,"Nº clientes":fmt_int})

    # R$/kg
    rkg_lin, rkg_cat, _, _ = preco_kg_por_linha(vendas_f)
    if not rkg_lin.empty:
        add_aba("R$ por kg (Linha)", rkg_lin, {
            "Faturamento":fmt_brl,"Peso_kg":fmt_int,"Quantidade":fmt_int,"R$/kg":fmt_rkg})
    if not rkg_cat.empty:
        add_aba("R$ por kg (Categoria)", rkg_cat, {
            "Faturamento":fmt_brl,"Peso_kg":fmt_int,"R$/kg":fmt_rkg})

    # Matéria-Prima
    mp_base = vendas_f[vendas_f["Materia_Prima"].isin(["Úmido","Integral"])]
    if not mp_base.empty:
        mp = mp_base.groupby("Materia_Prima", as_index=False).agg(
            Faturamento=("Vlr.Total","sum"), Quantidade=("Quantidade","sum"),
            Peso_kg=("Peso_kg_total","sum"))
        add_aba("Matéria-Prima", mp, {"Faturamento":fmt_brl,"Quantidade":fmt_int,"Peso_kg":fmt_int})

    wb.save(out); out.seek(0)
    return out.getvalue()


# =============================================================================
# EXPORTAÇÃO PDF
# =============================================================================
def _mpl_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight")
    plt.close(fig); buf.seek(0)
    return buf

def gerar_pdf(vendas_f, devol_f, filtros_desc):
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4,
                            leftMargin=1.3*cm, rightMargin=1.3*cm,
                            topMargin=1.3*cm, bottomMargin=1.3*cm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Heading1"], fontSize=17,
                              textColor=colors.HexColor("#102A43"), spaceAfter=8)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12,
                        textColor=colors.HexColor("#0B5FFF"), spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=9.5)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8,
                           textColor=colors.HexColor("#627D98"))

    def tabela_kpi_pareada(data, cor_header="#0B5FFF"):
        t = Table(data, colWidths=[6*cm, 4*cm, 6*cm, 4*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor(cor_header)),
            ("TEXTCOLOR",(0,0),(-1,0), colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1), 9),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("ALIGN",(3,0),(3,-1),"RIGHT"),
            ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
        ]))
        return t

    def tabela_padrao(data, col_widths, header_color="#0B5FFF", alt="#F0F4F8"):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor(header_color)),
            ("TEXTCOLOR",(0,0),(-1,0), colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1), 8.5),
            ("ALIGN",(1,0),(-1,-1),"RIGHT"),
            ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor(alt)]),
        ]))
        return t

    story = []
    # =========== PÁGINA 1: DIAGNÓSTICO EXECUTIVO ===========
    story.append(Paragraph("Diagnóstico Executivo — Coco & Cia", title_st))
    story.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Filtros: <b>{filtros_desc or 'Nenhum (consolidado)'}</b>", small))
    story.append(Spacer(1, 6))

    kpi = calc_kpis(vendas_f, devol_f)
    story.append(Paragraph("Indicadores-Chave", h2))
    kpi_data = [
        ["Faturamento Bruto", fmt_brl(kpi["bruto"]),
         "Quantidade Vendida (CX)", fmt_int(kpi["qtd_v"])],
        ["Devoluções (R$)", fmt_brl(kpi["devol_vlr"]),
         "Peso Vendido (kg)", fmt_int(kpi["peso_kg"])],
        ["Faturamento Líquido", fmt_brl(kpi["liquido"]),
         "Preço Médio por CX", fmt_brl(kpi["preco_medio_unit"])],
        ["% Devolução", fmt_pct(kpi["pct_devol"]),
         "Preço Médio por kg", fmt_rkg(kpi["preco_medio_kg"])],
        ["Clientes Ativos", fmt_int(kpi["clientes_ativos"]),
         "SKUs Vendidos", fmt_int(kpi["produtos_vendidos"])],
    ]
    story.append(tabela_kpi_pareada(kpi_data))
    story.append(Spacer(1, 8))

    # Gráfico mensal
    mv = vendas_f.groupby("AnoMes")["Vlr.Total"].sum()
    md = devol_f.groupby("AnoMes")["Vlr.Total"].sum()
    mensal = pd.DataFrame({"V":mv,"D":md}).fillna(0).sort_index()
    mensal["L"] = mensal["V"] - mensal["D"]
    if not mensal.empty:
        story.append(Paragraph("Evolução Mensal", h2))
        fig, ax = plt.subplots(figsize=(8, 2.6))
        x = mensal.index.strftime("%b/%y")
        ax.bar(x, mensal["V"], color="#0B5FFF", label="Bruto")
        ax.bar(x, -mensal["D"], color="#D64545", label="Devoluções")
        ax.plot(x, mensal["L"], color="#27AB83", marker="o", linewidth=1.8, label="Líquido")
        ax.set_ylabel("R$"); ax.legend(loc="upper left", fontsize=7)
        ax.tick_params(axis="x", rotation=45, labelsize=7); ax.grid(axis="y", alpha=0.3)
        story.append(RLImage(_mpl_to_bytes(fig), width=18*cm, height=6*cm))

    # Resumo YoY (lado a lado com top alertas)
    if {2025, 2026}.issubset(set(vendas_f["Ano"].unique())):
        story.append(Paragraph("Comparativo YoY (meses comuns)", h2))
        comuns = sorted(set(vendas_f[vendas_f["Ano"]==2025]["Mes"]) & set(vendas_f[vendas_f["Ano"]==2026]["Mes"]))
        v25 = vendas_f[(vendas_f["Ano"]==2025)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        v26 = vendas_f[(vendas_f["Ano"]==2026)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        d25 = devol_f[(devol_f["Ano"]==2025)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        d26 = devol_f[(devol_f["Ano"]==2026)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
        yoy_data = [["Indicador","2025","2026","Δ Abs","Δ %"]]
        for nome, a, b in [("Bruto", v25, v26), ("Devoluções", d25, d26), ("Líquido", v25-d25, v26-d26)]:
            yoy_data.append([nome, fmt_brl(a), fmt_brl(b), fmt_brl(b-a),
                              fmt_delta_pct((b-a)/a) if a else "—"])
        story.append(tabela_padrao(yoy_data, [3.6*cm,3.6*cm,3.6*cm,3.6*cm,2.6*cm]))

    story.append(PageBreak())

    # =========== PÁGINA 2: CURVA ABC ===========
    abc, total = analise_abc(vendas_f, devol_f)
    if not abc.empty and total > 0:
        story.append(Paragraph("Curva ABC de Produtos", h2))
        cresumo = abc.groupby("Classe").agg(qtd=("Produto","count"), liq=("Líquido","sum")).reset_index()
        cresumo["pct"] = cresumo["liq"] / total
        data = [["Classe","Qtd Produtos","Líquido","% Faturamento"]]
        for _, r in cresumo.iterrows():
            data.append([r["Classe"], fmt_int(r["qtd"]), fmt_brl(r["liq"]), fmt_pct(r["pct"])])
        story.append(tabela_padrao(data, [3*cm,4*cm,5*cm,4*cm]))
        story.append(Spacer(1, 6))

        story.append(Paragraph("Top 15 Produtos (Classe A)", body))
        top15 = abc.head(15)
        data = [["Produto","Líquido","% Acum.","Classe"]]
        for _, r in top15.iterrows():
            data.append([str(r["Produto_Label"])[:50], fmt_brl(r["Líquido"]),
                          fmt_pct(r["% Acumulado"]), r["Classe"]])
        story.append(tabela_padrao(data, [9*cm,3.5*cm,3*cm,1.8*cm]))

    story.append(PageBreak())

    # =========== PÁGINA 3: R$/kg + Preço Médio ===========
    rkg_lin, rkg_cat, alerta, texto_alerta = preco_kg_por_linha(vendas_f)
    if not rkg_lin.empty:
        story.append(Paragraph("Análise R$/kg por Linha", h2))
        story.append(Paragraph(
            "Peso extraído da descrição do SKU. Em tese, embalagens maiores "
            "(Profissional 5kg) devem ter R$/kg menor que embalagens menores (Food 10×1kg). "
            "Desvio dessa lógica indica distorção de pricing.", small))
        data = [["Linha","Faturamento","Peso (kg)","R$/kg"]]
        for _, r in rkg_lin.iterrows():
            data.append([r["Linha"], fmt_brl(r["Faturamento"]),
                          fmt_int(r["Peso_kg"]), fmt_rkg(r["R$/kg"])])
        story.append(tabela_padrao(data, [4*cm,4.5*cm,3*cm,3.5*cm]))
        if alerta:
            story.append(Spacer(1, 4))
            alerta_st = ParagraphStyle("alerta", parent=body, textColor=colors.HexColor("#7B1818"),
                                       backColor=colors.HexColor("#FEF2F2"), borderPadding=6)
            story.append(Paragraph(texto_alerta, alerta_st))
        story.append(Spacer(1, 6))

    pm = preco_medio_por_produto(vendas_f)
    if not pm.empty:
        story.append(Paragraph("Top 10 Produtos com Maior Amplitude de Preço entre Clientes", h2))
        story.append(Paragraph(
            "Amplitude = (Maior - Menor) / Menor preço cobrado por cliente. "
            "Valor alto indica clientes pagando preços muito diferentes pelo mesmo SKU.", small))
        pm_top = pm[pm["N"] >= 2].nlargest(10, "Amplitude_%")
        data = [["Produto","Preço Médio","Mín. cliente","Máx. cliente","Amplitude"]]
        for _, r in pm_top.iterrows():
            data.append([str(r["Produto_Label"])[:35], fmt_brl(r["Preço_Medio_CX"]),
                          fmt_brl(r["Min"]), fmt_brl(r["Max"]), fmt_pct(r["Amplitude_%"])])
        story.append(tabela_padrao(data, [6*cm,3.2*cm,2.8*cm,2.8*cm,2.4*cm]))

    story.append(PageBreak())

    # =========== PÁGINA 4: RISCO + DEVOL ===========
    risco = clientes_em_risco(vendas_f)
    if not risco.empty:
        story.append(Paragraph("Top 12 Clientes em Risco (queda ≥30% vs 2025)", h2))
        data = [["Cliente","2025","2026","Δ Abs","Nível"]]
        for _, r in risco.head(12).iterrows():
            data.append([str(r["Cliente_Label"])[:30], fmt_brl(r[2025]),
                          fmt_brl(r[2026]), fmt_brl(r["Δ Abs"]),
                          r["Nível"].split(" ", 1)[-1]])
        story.append(tabela_padrao(data, [6*cm,3*cm,3*cm,3*cm,2.5*cm], header_color="#D64545", alt="#FEF2F2"))
        story.append(Spacer(1, 8))

    pp, _ = devolucoes_problematicas(vendas_f, devol_f)
    if not pp.empty:
        story.append(Paragraph("Top 10 Produtos com Maior Taxa de Devolução", h2))
        data = [["Produto","Vendido","Devolvido","Taxa"]]
        for _, r in pp.head(10).iterrows():
            data.append([str(r["Produto_Label"])[:40], fmt_brl(r["Vendido"]),
                          fmt_brl(r["Devolvido"]), fmt_pct(r["Taxa"])])
        story.append(tabela_padrao(data, [8*cm,3.5*cm,3.5*cm,2.5*cm], header_color="#D64545", alt="#FEF2F2"))

    # Geo
    gr = crescimento_geo(vendas_f, devol_f, "Regiao_Desc")
    if not gr.empty:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Crescimento por Região (Líquido, meses comuns)", h2))
        data = [["Região","2025","2026","Δ Abs","Δ %"]]
        for _, r in gr.iterrows():
            data.append([str(r["Regiao_Desc"])[:25], fmt_brl(r[2025]),
                          fmt_brl(r[2026]), fmt_brl(r["Δ Abs"]),
                          fmt_delta_pct(r["Δ %"])])
        story.append(tabela_padrao(data, [4*cm,3.5*cm,3.5*cm,3.5*cm,2.6*cm]))

    # Rodapé
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        "Devoluções abatidas no mês da DT Digitacao (regra contábil). "
        "Peso por kg extraído da descrição do SKU; líquidos tratados com densidade ≈1.",
        small))

    doc.build(story); out.seek(0)
    return out.getvalue()


# =============================================================================
# UI — SIDEBAR (UPLOAD + FILTROS CASCATA)
# =============================================================================
st.sidebar.title("📁 Upload de Bases")
up_vendas = st.sidebar.file_uploader("Vendas (.xlsx)", type=["xlsx","xls"], key="up_v")
up_devol  = st.sidebar.file_uploader("Devoluções (.xlsx)", type=["xlsx","xls"], key="up_d")
up_cli    = st.sidebar.file_uploader("Banco de Clientes (.xlsx)", type=["xlsx","xls"], key="up_c")

if not (up_vendas and up_devol and up_cli):
    st.title("📊 Dashboard Comercial — Coco & Cia")
    st.markdown(
        """
        Carregue os **3 arquivos** na barra lateral para começar:

        1. **Base de Vendas** — `Emissao`, `Cliente`, `Produto`, `Quantidade`, `Vlr.Total`
        2. **Base de Devoluções** — `DT Digitacao`, `Forn/Cliente`, `Produto`, `Quantidade`, `Vlr.Total`
        3. **Banco de Clientes** — `Codigo`, `Nome`, `Regiao`, `Estado`, `Municipio`

        > ⚠️ Devoluções são abatidas **estritamente no mês da `DT Digitacao`**.

        **Recursos desta versão:**
        - 🎯 Diagnóstico executivo em uma página
        - ⚖️ Comparação entre clientes ou estados (lado a lado)
        - 💰 Preço médio por cliente e detecção de distorção entre concorrentes
        - ⚖️ R$/kg por linha com alerta de contra-senso (Food vs Profissional)
        - 🎯 Curva ABC de produtos
        - ⚠️ Clientes em risco, devoluções problemáticas, ranking de vendedores
        - 📥 Exportação completa em **Excel** e **PDF**
        - 🎯 Filtros em cascata (Região → Estado → Município)
        """
    )
    st.stop()

try:
    vendas_raw = load_vendas(up_vendas.getvalue())
    devol_raw  = load_devolucoes(up_devol.getvalue())
    clientes   = load_clientes(up_cli.getvalue())
except Exception as e:
    st.error(f"Erro ao processar arquivos: {e}")
    st.stop()

vendas = enrich(vendas_raw, clientes)
devol  = enrich(devol_raw, clientes)
universe = pd.concat([vendas, devol], ignore_index=True)

# ----- FILTROS CASCATA -----
st.sidebar.markdown("---")
st.sidebar.title("🎯 Filtros (Cascata)")
st.sidebar.caption("Cada filtro limita as opções dos seguintes.")

opcoes_reg = sorted([x for x in universe["Regiao_Desc"].dropna().unique() if str(x) != "nan"])
f_regiao = st.sidebar.multiselect("🌎 Região", opcoes_reg, default=[])
s1 = universe if not f_regiao else universe[universe["Regiao_Desc"].isin(f_regiao)]

opcoes_uf = sorted([x for x in s1["Estado"].dropna().unique() if str(x) != "nan"])
f_estado = st.sidebar.multiselect("📍 Estado (UF)", opcoes_uf, default=[])
s2 = s1 if not f_estado else s1[s1["Estado"].isin(f_estado)]

opcoes_mun = sorted([x for x in s2["Municipio"].dropna().unique() if str(x) != "nan"])
f_municipio = st.sidebar.multiselect("🏙️ Município", opcoes_mun, default=[])
s3 = s2 if not f_municipio else s2[s2["Municipio"].isin(f_municipio)]

opcoes_cli = sorted([x for x in s3["Cliente_Label"].dropna().unique() if str(x) != "nan"])
f_cliente = st.sidebar.multiselect("👤 Cliente (código — nome)", opcoes_cli, default=[],
                                    help="Apenas clientes da seleção geográfica acima.")
s4 = s3 if not f_cliente else s3[s3["Cliente_Label"].isin(f_cliente)]

opcoes_linha = sorted(s4["Linha"].dropna().unique())
f_linha = st.sidebar.multiselect("📦 Linha de Atuação", opcoes_linha, default=[])
s5 = s4 if not f_linha else s4[s4["Linha"].isin(f_linha)]

opcoes_cat = sorted(s5["Categoria"].dropna().unique())
f_categoria = st.sidebar.multiselect("🏷️ Categoria", opcoes_cat, default=[])

def aplicar(df):
    out = df
    if f_regiao: out = out[out["Regiao_Desc"].isin(f_regiao)]
    if f_estado: out = out[out["Estado"].isin(f_estado)]
    if f_municipio: out = out[out["Municipio"].isin(f_municipio)]
    if f_cliente: out = out[out["Cliente_Label"].isin(f_cliente)]
    if f_linha: out = out[out["Linha"].isin(f_linha)]
    if f_categoria: out = out[out["Categoria"].isin(f_categoria)]
    return out

vendas_f = aplicar(vendas)
devol_f  = aplicar(devol)

filtros_lista = []
if f_regiao: filtros_lista.append(f"Região: {', '.join(f_regiao)}")
if f_estado: filtros_lista.append(f"UF: {', '.join(f_estado)}")
if f_municipio: filtros_lista.append(f"Município: {len(f_municipio)} selec.")
if f_cliente: filtros_lista.append(f"Cliente: {len(f_cliente)} selec.")
if f_linha: filtros_lista.append(f"Linha: {', '.join(f_linha)}")
if f_categoria: filtros_lista.append(f"Categoria: {', '.join(f_categoria)}")
filtros_desc = " | ".join(filtros_lista)

st.sidebar.markdown("---")
st.sidebar.caption(f"📊 Vendas: **{fmt_int(len(vendas_f))}** linhas")
st.sidebar.caption(f"↩️ Devol.: **{fmt_int(len(devol_f))}** linhas")

# ----- EXPORTAÇÃO -----
st.sidebar.markdown("---")
st.sidebar.title("📥 Exportar Relatório")
st.sidebar.caption("Gera com os filtros aplicados.")

col_e1, col_e2 = st.sidebar.columns(2)
if not vendas_f.empty:
    with col_e1:
        excel_bytes = gerar_excel(vendas_f, devol_f, filtros_desc)
        st.download_button("📊 Excel", excel_bytes,
            file_name=f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with col_e2:
        pdf_bytes = gerar_pdf(vendas_f, devol_f, filtros_desc)
        st.download_button("📄 PDF", pdf_bytes,
            file_name=f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True)


# =============================================================================
# CABEÇALHO E KPIs PRINCIPAIS (visível em todas as abas)
# =============================================================================
st.title("📊 Dashboard Comercial — Coco & Cia")
st.caption("Devoluções abatidas pelo mês da `DT Digitacao` · "
            "Peso/kg extraído da descrição do SKU · "
            "Cliente e Produto exibidos como `Código — Nome`.")

kpi = calc_kpis(vendas_f, devol_f)
k1, k2, k3, k4, k5, k6, k7, k8 = st.columns(8)
k1.metric("💰 Bruto", fmt_brl(kpi["bruto"]))
k2.metric("✅ Líquido", fmt_brl(kpi["liquido"]),
          delta=f"-{fmt_pct(kpi['pct_devol'])} devol.", delta_color="inverse")
k3.metric("📦 Qtd (CX)", fmt_int(kpi["qtd_v"]))
k4.metric("↩️ Devol. R$", fmt_brl(kpi["devol_vlr"]))
k5.metric("👥 Clientes Ativos", fmt_int(kpi["clientes_ativos"]))
k6.metric("⚖️ Peso (kg)", fmt_int(kpi["peso_kg"]))
k7.metric("💵 Preço/CX", fmt_brl(kpi["preco_medio_unit"]))
k8.metric("💵 Preço/kg", fmt_rkg(kpi["preco_medio_kg"]))


# =============================================================================
# TABS
# =============================================================================
tab_diag, tab_geral, tab_yoy, tab_mp, tab_preco, tab_comp, tab_estrat, tab_risco, tab_dados = st.tabs([
    "🎯 Diagnóstico", "📈 Visão Geral", "📅 YoY",
    "🥥 Matéria-Prima & R$/kg", "💰 Preço Médio",
    "⚖️ Comparação", "🔍 Estratégicas", "⚠️ Riscos", "🗂️ Dados",
])

# -----------------------------------------------------------------------------
# TAB 1 — DIAGNÓSTICO EXECUTIVO (TUDO EM UMA PÁGINA)
# -----------------------------------------------------------------------------
with tab_diag:
    if vendas_f.empty:
        st.warning("Sem dados para os filtros atuais.")
    else:
        st.markdown(
            "Diagnóstico consolidado em uma página. "
            "Para detalhes navegue pelas abas; para imprimir, **exporte o PDF** na barra lateral."
        )

        # ---- KPIs adicionais ----
        kk1, kk2, kk3, kk4 = st.columns(4)
        kk1.metric("👥 Clientes Ativos", fmt_int(kpi["clientes_ativos"]))
        kk2.metric("🏷️ SKUs Vendidos", fmt_int(kpi["produtos_vendidos"]))
        kk3.metric("↩️ Devol. R$", fmt_brl(kpi["devol_vlr"]))
        kk4.metric("↩️ % Devol.", fmt_pct(kpi["pct_devol"]))

        st.markdown("---")
        # ---- ALERTAS (curto) ----
        st.subheader("⚠️ Alertas-Chave")
        alertas = []

        # Alerta 1: R$/kg contra-senso
        _, _, alerta_rkg, txt_rkg = preco_kg_por_linha(vendas_f)
        if alerta_rkg:
            alertas.append(("danger", txt_rkg))

        # Alerta 2: Clientes em risco
        risco_df = clientes_em_risco(vendas_f)
        if not risco_df.empty:
            n_crit = (risco_df["Nível"] == "🔴 Crítico (parou)").sum()
            if n_crit > 0:
                alertas.append(("danger", f"🔴 {n_crit} cliente(s) crítico(s) — pararam de comprar em 2026 (vs 2025)."))
            n_sev = (risco_df["Nível"] == "🔴 Severo").sum()
            if n_sev > 0:
                alertas.append(("warn", f"🟠 {n_sev} cliente(s) com queda >70% vs 2025."))

        # Alerta 3: YoY
        if {2025, 2026}.issubset(set(vendas_f["Ano"].unique())):
            comuns = sorted(set(vendas_f[vendas_f["Ano"]==2025]["Mes"]) & set(vendas_f[vendas_f["Ano"]==2026]["Mes"]))
            v25 = vendas_f[(vendas_f["Ano"]==2025)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
            v26 = vendas_f[(vendas_f["Ano"]==2026)&(vendas_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
            d25 = devol_f[(devol_f["Ano"]==2025)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
            d26 = devol_f[(devol_f["Ano"]==2026)&(devol_f["Mes"].isin(comuns))]["Vlr.Total"].sum()
            l25 = v25 - d25; l26 = v26 - d26
            if l25 > 0:
                delta = (l26 - l25) / l25
                if delta <= -0.05:
                    alertas.append(("danger",
                        f"📉 Faturamento líquido caiu **{fmt_delta_pct(delta)}** em 2026 vs mesmos meses de 2025."))
                elif delta >= 0.05:
                    alertas.append(("ok",
                        f"📈 Faturamento líquido cresceu **{fmt_delta_pct(delta)}** em 2026 vs mesmos meses de 2025."))

        # Alerta 4: amplitude de preço alta
        pm_df = preco_medio_por_produto(vendas_f)
        if not pm_df.empty:
            distorcao = pm_df[(pm_df["N"]>=3) & (pm_df["Amplitude_%"] > 0.5)]
            if not distorcao.empty:
                alertas.append(("warn",
                    f"💰 {len(distorcao)} produto(s) com amplitude de preço >50% entre clientes — possível distorção de pricing."))

        if not alertas:
            st.markdown('<div class="alert-box alert-ok">✅ Nenhum alerta crítico nos filtros atuais.</div>', unsafe_allow_html=True)
        else:
            for tipo, msg in alertas[:6]:
                cls = {"danger":"alert-danger", "warn":"alert-warn", "ok":"alert-ok"}[tipo]
                st.markdown(f'<div class="alert-box {cls}">{msg}</div>', unsafe_allow_html=True)

        st.markdown("---")
        # ---- LINHA: mensal + mix linha ----
        ld1, ld2 = st.columns([2, 1])
        with ld1:
            st.subheader("Evolução Mensal")
            mv = vendas_f.groupby("AnoMes", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"V"})
            md = devol_f.groupby("AnoMes", as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"D"})
            mens = mv.merge(md, on="AnoMes", how="outer").fillna(0).sort_values("AnoMes")
            mens["L"] = mens["V"] - mens["D"]
            mens["Label"] = mens["AnoMes"].dt.strftime("%b/%y")
            fig = go.Figure()
            fig.add_bar(x=mens["Label"], y=mens["V"], name="Bruto", marker_color="#0B5FFF")
            fig.add_bar(x=mens["Label"], y=-mens["D"], name="Devol.", marker_color="#D64545")
            fig.add_trace(go.Scatter(x=mens["Label"], y=mens["L"], mode="lines+markers",
                                       name="Líquido", line=dict(color="#27AB83", width=2.5)))
            fig.update_layout(barmode="relative", height=300, hovermode="x unified",
                              margin=dict(l=10,r=10,t=10,b=10),
                              legend=dict(orientation="h", y=1.1, x=1, xanchor="right"))
            st.plotly_chart(fig, use_container_width=True)
        with ld2:
            st.subheader("Mix por Linha")
            mx = vendas_f.groupby("Linha", as_index=False)["Vlr.Total"].sum()
            fig_p = px.pie(mx, names="Linha", values="Vlr.Total", hole=0.55,
                            color="Linha", color_discrete_map=COLORS)
            fig_p.update_traces(textposition="outside", textinfo="label+percent")
            fig_p.update_layout(height=300, margin=dict(l=10,r=10,t=10,b=10), showlegend=False)
            st.plotly_chart(fig_p, use_container_width=True)

        # ---- LINHA: top 5 produtos, top 5 clientes ----
        ld3, ld4 = st.columns(2)
        with ld3:
            st.subheader("🏆 Top 5 Produtos (Líquido)")
            abc, _ = analise_abc(vendas_f, devol_f)
            if not abc.empty:
                t5p = abc.head(5)[["Produto_Label","Líquido","Classe"]]
                t5p.columns = ["Produto", "Líquido", "Classe"]
                st.dataframe(t5p.style.format({"Líquido": fmt_brl}),
                              use_container_width=True, hide_index=True)
        with ld4:
            st.subheader("🏆 Top 5 Clientes (Líquido)")
            tcv = vendas_f.groupby(["Cliente","Nome"], as_index=False)["Vlr.Total"].sum()
            tcd = devol_f.groupby("Cliente", as_index=False)["Vlr.Total"].sum()
            tc = tcv.merge(tcd, on="Cliente", how="left", suffixes=("_v","_d")).fillna(0)
            tc["Líquido"] = tc["Vlr.Total_v"] - tc["Vlr.Total_d"]
            tc["Cliente"] = tc["Cliente"].astype(str) + " — " + tc["Nome"].astype(str)
            t5 = tc.nlargest(5, "Líquido")[["Cliente","Líquido"]]
            st.dataframe(t5.style.format({"Líquido": fmt_brl}),
                          use_container_width=True, hide_index=True)

# -----------------------------------------------------------------------------
# TAB 2 — VISÃO GERAL
# -----------------------------------------------------------------------------
with tab_geral:
    if vendas_f.empty:
        st.warning("Sem vendas para os filtros selecionados.")
    else:
        mv = vendas_f.groupby("AnoMes", as_index=False).agg(Vendas=("Vlr.Total","sum"))
        md = devol_f.groupby("AnoMes", as_index=False).agg(Devolucoes=("Vlr.Total","sum"))
        mensal = mv.merge(md, on="AnoMes", how="outer").fillna(0).sort_values("AnoMes")
        mensal["Liquido"] = mensal["Vendas"] - mensal["Devolucoes"]
        mensal["Label"] = mensal["AnoMes"].dt.strftime("%b/%y")

        st.subheader("Evolução Mensal — Bruto, Devoluções e Líquido")
        fig = go.Figure()
        fig.add_bar(x=mensal["Label"], y=mensal["Vendas"], name="Bruto", marker_color=COLORS["Venda"])
        fig.add_bar(x=mensal["Label"], y=-mensal["Devolucoes"], name="Devoluções", marker_color=COLORS["Devolução"])
        fig.add_trace(go.Scatter(x=mensal["Label"], y=mensal["Liquido"], mode="lines+markers",
                                   name="Líquido", line=dict(color=COLORS["Líquido"], width=3),
                                   marker=dict(size=8)))
        fig.update_layout(barmode="relative", height=420, hovermode="x unified",
                           yaxis_title="R$", xaxis_title="",
                           legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                           margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig, use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Mix por Linha")
            ml = vendas_f.groupby("Linha", as_index=False)["Vlr.Total"].sum().sort_values("Vlr.Total", ascending=False)
            fig_l = px.pie(ml, names="Linha", values="Vlr.Total", hole=0.55,
                            color="Linha", color_discrete_map=COLORS)
            fig_l.update_traces(textposition="outside", textinfo="label+percent")
            fig_l.update_layout(height=350, margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_l, use_container_width=True)
        with c2:
            st.subheader("Faturamento por Categoria")
            mc = vendas_f.groupby("Categoria", as_index=False)["Vlr.Total"].sum().sort_values("Vlr.Total")
            fig_c = px.bar(mc, x="Vlr.Total", y="Categoria", orientation="h",
                            text=mc["Vlr.Total"].apply(fmt_brl),
                            color="Vlr.Total", color_continuous_scale="Blues")
            fig_c.update_traces(textposition="outside")
            fig_c.update_layout(height=350, showlegend=False, coloraxis_showscale=False,
                                 yaxis_title="", xaxis_title="R$", margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_c, use_container_width=True)

        c3, c4 = st.columns(2)
        with c3:
            st.subheader("Top 10 Produtos (Líquido)")
            tv = vendas_f.groupby(["Produto","Descricao"], as_index=False)["Vlr.Total"].sum()
            td = devol_f.groupby("Produto", as_index=False)["Vlr.Total"].sum()
            tp = tv.merge(td, on="Produto", how="left", suffixes=("_v","_d")).fillna(0)
            tp["Líquido"] = tp["Vlr.Total_v"] - tp["Vlr.Total_d"]
            tp = tp.nlargest(10, "Líquido")
            tp["Produto"] = tp["Produto"].astype(str) + " — " + tp["Descricao"].astype(str)
            tp = tp[["Produto","Vlr.Total_v","Vlr.Total_d","Líquido"]]
            tp.columns = ["Produto","Bruto","Devoluções","Líquido"]
            st.dataframe(tp.style.format({"Bruto":fmt_brl,"Devoluções":fmt_brl,"Líquido":fmt_brl}),
                          use_container_width=True, hide_index=True)
        with c4:
            st.subheader("Top 10 Clientes (Líquido)")
            tcv = vendas_f.groupby(["Cliente","Nome"], as_index=False)["Vlr.Total"].sum()
            tcd = devol_f.groupby("Cliente", as_index=False)["Vlr.Total"].sum()
            tc = tcv.merge(tcd, on="Cliente", how="left", suffixes=("_v","_d")).fillna(0)
            tc["Líquido"] = tc["Vlr.Total_v"] - tc["Vlr.Total_d"]
            tc = tc.nlargest(10, "Líquido")
            tc["Cliente"] = tc["Cliente"].astype(str) + " — " + tc["Nome"].astype(str)
            tc = tc[["Cliente","Vlr.Total_v","Vlr.Total_d","Líquido"]]
            tc.columns = ["Cliente","Bruto","Devoluções","Líquido"]
            st.dataframe(tc.style.format({"Bruto":fmt_brl,"Devoluções":fmt_brl,"Líquido":fmt_brl}),
                          use_container_width=True, hide_index=True)

# -----------------------------------------------------------------------------
# TAB 3 — YOY
# -----------------------------------------------------------------------------
with tab_yoy:
    anos_disp = sorted(vendas_f["Ano"].dropna().unique().tolist())
    anos_target = [a for a in (2025, 2026) if a in anos_disp]
    if len(anos_target) < 2:
        st.info(f"Para o YoY 2025 vs 2026, a base precisa ter ambos os anos. Atual: {anos_target or 'nenhum'}.")
    else:
        vy = vendas_f[vendas_f["Ano"].isin(anos_target)].groupby(["Ano","Mes"], as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"V"})
        dy = devol_f[devol_f["Ano"].isin(anos_target)].groupby(["Ano","Mes"], as_index=False)["Vlr.Total"].sum().rename(columns={"Vlr.Total":"D"})
        yoy = vy.merge(dy, on=["Ano","Mes"], how="outer").fillna(0)
        yoy["Líquido"] = yoy["V"] - yoy["D"]
        yoy["Mes_Lbl"] = yoy["Mes"].map(MESES_PT)
        yoy["Mes_Lbl"] = pd.Categorical(yoy["Mes_Lbl"], categories=list(MESES_PT.values()), ordered=True)
        yoy = yoy.sort_values(["Ano","Mes"])

        st.subheader("Faturamento Líquido Mensal — 2025 vs 2026")
        fig_yoy = px.line(yoy, x="Mes_Lbl", y="Líquido", color="Ano", markers=True,
                           color_discrete_map={2025: COLORS["2025"], 2026: COLORS["2026"]})
        fig_yoy.update_traces(line=dict(width=3), marker=dict(size=9))
        fig_yoy.update_layout(height=400, xaxis_title="", yaxis_title="R$ Líquido",
                               hovermode="x unified", legend=dict(title="Ano"),
                               margin=dict(l=10,r=10,t=20,b=10))
        st.plotly_chart(fig_yoy, use_container_width=True)

        comuns = sorted(set(yoy[yoy["Ano"]==2025]["Mes"]) & set(yoy[yoy["Ano"]==2026]["Mes"]))
        base = yoy[yoy["Mes"].isin(comuns)]
        tot25 = base[base["Ano"]==2025][["V","D","Líquido"]].sum()
        tot26 = base[base["Ano"]==2026][["V","D","Líquido"]].sum()

        st.subheader(f"Resumo Comparável — meses: {', '.join(MESES_PT[m] for m in comuns)}")
        comp = pd.DataFrame({
            "Indicador":["Faturamento Bruto","Devoluções (R$)","Faturamento Líquido"],
            "2025":[tot25["V"], tot25["D"], tot25["Líquido"]],
            "2026":[tot26["V"], tot26["D"], tot26["Líquido"]]})
        comp["Δ Abs"] = comp["2026"] - comp["2025"]
        comp["Δ %"] = (comp["Δ Abs"]) / comp["2025"].replace(0, pd.NA)
        st.dataframe(comp.style.format({"2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct}),
                      use_container_width=True, hide_index=True)

        st.subheader("Performance por Categoria — 2025 vs 2026 (meses comuns)")
        bv = vendas_f[vendas_f["Mes"].isin(comuns)]
        bd = devol_f[devol_f["Mes"].isin(comuns)]
        cv = bv.groupby(["Categoria","Ano"])["Vlr.Total"].sum().unstack(fill_value=0)
        cd = bd.groupby(["Categoria","Ano"])["Vlr.Total"].sum().unstack(fill_value=0)
        for a in (2025, 2026):
            if a not in cv.columns: cv[a] = 0
            if a not in cd.columns: cd[a] = 0
        cl = (cv - cd).reset_index()
        fig_cy = go.Figure()
        for ano, cor in [(2025, COLORS["2025"]), (2026, COLORS["2026"])]:
            fig_cy.add_bar(x=cl["Categoria"], y=cl[ano], name=str(ano), marker_color=cor,
                             text=[fmt_brl(v) for v in cl[ano]], textposition="outside")
        fig_cy.update_layout(barmode="group", height=420, yaxis_title="R$ Líquido",
                               hovermode="x unified", margin=dict(l=10,r=10,t=20,b=10))
        st.plotly_chart(fig_cy, use_container_width=True)


# -----------------------------------------------------------------------------
# TAB 4 — MATÉRIA-PRIMA & R$/KG
# -----------------------------------------------------------------------------
with tab_mp:
    st.markdown("Decisões baseadas em **matéria-prima** e **rentabilidade por kg** das linhas.")

    # ----- Bloco 1: Úmido vs Integral -----
    base_mp = vendas_f[vendas_f["Materia_Prima"].isin(["Úmido","Integral"])]
    if not base_mp.empty:
        st.subheader("🥥 Úmido vs Integral (Escassez de Matéria-Prima)")
        m1, m2, m3 = st.columns(3)
        q_umi = base_mp[base_mp["Materia_Prima"]=="Úmido"]["Quantidade"].sum()
        q_int = base_mp[base_mp["Materia_Prima"]=="Integral"]["Quantidade"].sum()
        total = q_umi + q_int
        m1.metric("Qtd Úmido (CX)", fmt_int(q_umi), delta=fmt_pct(q_umi/total if total else 0))
        m2.metric("Qtd Integral (CX)", fmt_int(q_int), delta=fmt_pct(q_int/total if total else 0))
        m3.metric("Ratio Úmido/Integral", f"{(q_umi/q_int):.2f}x" if q_int else "—")

        mp_m = base_mp.groupby(["AnoMes","Materia_Prima"], as_index=False)["Quantidade"].sum()
        mp_m["Label"] = mp_m["AnoMes"].dt.strftime("%b/%y")
        fig_mp = px.bar(mp_m, x="Label", y="Quantidade", color="Materia_Prima",
                          barmode="group", color_discrete_map=COLORS)
        fig_mp.update_layout(height=350, hovermode="x unified",
                              xaxis_title="", yaxis_title="Quantidade (CX)",
                              margin=dict(l=10,r=10,t=20,b=10))
        st.plotly_chart(fig_mp, use_container_width=True)

        st.subheader("Mix Mensal (%)")
        mp_share = mp_m.pivot(index="Label", columns="Materia_Prima",
                               values="Quantidade").fillna(0)
        mp_share = mp_share.div(mp_share.sum(axis=1).replace(0, pd.NA), axis=0) * 100
        mp_share = mp_share.reset_index().melt(id_vars="Label",
                                                 var_name="Materia_Prima", value_name="Share")
        ordem = mp_m.sort_values("AnoMes")["Label"].unique().tolist()
        mp_share["Label"] = pd.Categorical(mp_share["Label"], categories=ordem, ordered=True)
        fig_s = px.bar(mp_share.sort_values("Label"), x="Label", y="Share",
                        color="Materia_Prima", color_discrete_map=COLORS)
        fig_s.update_layout(barmode="stack", height=320, yaxis_title="% do volume",
                             xaxis_title="", margin=dict(l=10,r=10,t=20,b=10))
        st.plotly_chart(fig_s, use_container_width=True)

    st.markdown("---")
    # ----- Bloco 2: R$/kg por Linha -----
    st.subheader("⚖️ R$/kg por Linha (Pricing Consistency Check)")
    st.caption(
        "Peso da embalagem extraído da descrição do SKU. "
        "Em tese, **embalagem maior** (Profissional 5kg) **deveria ter R$/kg menor** que embalagem menor (Food 10×1kg). "
        "Inversão dessa lógica = sinal de distorção de pricing."
    )
    rkg_lin, rkg_cat, alerta, txt_alerta = preco_kg_por_linha(vendas_f)

    if rkg_lin.empty:
        st.info("Sem produtos com peso extraível nos filtros atuais.")
    else:
        if alerta:
            st.error(txt_alerta)
        else:
            st.success("✅ Sem inversão de R$/kg entre Food e Profissional nos filtros atuais.")

        cols_rkg = st.columns(len(rkg_lin))
        for i, (_, r) in enumerate(rkg_lin.iterrows()):
            cor = COLORS.get(r["Linha"], "#000")
            cols_rkg[i].metric(
                f"Linha {r['Linha']}", fmt_rkg(r["R$/kg"]),
                delta=f"{fmt_brl(r['Faturamento'])} | {fmt_int(r['Peso_kg'])} kg",
                delta_color="off",
            )

        fig_rkg = px.bar(rkg_lin, x="Linha", y="R$/kg", color="Linha",
                          color_discrete_map=COLORS, text=rkg_lin["R$/kg"].apply(fmt_rkg))
        fig_rkg.update_traces(textposition="outside")
        fig_rkg.update_layout(height=330, showlegend=False, yaxis_title="R$ por kg",
                                margin=dict(l=10,r=10,t=20,b=10))
        st.plotly_chart(fig_rkg, use_container_width=True)

        st.subheader("R$/kg por Categoria (detalhe)")
        st.dataframe(rkg_cat.style.format({
            "Faturamento":fmt_brl,"Peso_kg":fmt_int,"R$/kg":fmt_rkg,
        }), use_container_width=True, hide_index=True)


# -----------------------------------------------------------------------------
# TAB 5 — PREÇO MÉDIO (DETECTAR DISTORÇÃO ENTRE CLIENTES)
# -----------------------------------------------------------------------------
with tab_preco:
    st.markdown(
        "**Detecção de distorção de preço entre clientes.** "
        "Para cada SKU, calcula o preço médio que cada cliente paga e mostra a amplitude."
    )

    pm = preco_medio_por_produto(vendas_f)
    if pm.empty:
        st.warning("Sem dados para análise de preço.")
    else:
        # ----- Top produtos com maior dispersão -----
        st.subheader("📊 Produtos com Maior Amplitude de Preço entre Clientes")
        st.caption("Amplitude = (preço máximo - preço mínimo) / preço mínimo. "
                    "Valores altos indicam que clientes diferentes pagam preços muito diferentes pelo mesmo SKU.")

        pm_show = pm[pm["N"] >= 2].copy()
        if pm_show.empty:
            st.info("Sem produtos vendidos para pelo menos 2 clientes diferentes.")
        else:
            tabela = pm_show[["Produto_Label","Faturamento","Preço_Medio_CX",
                                "Min","Max","Amplitude_%","N"]].head(20)
            tabela.columns = ["Produto","Faturamento","Preço Médio (CX)",
                               "Mín. entre clientes","Máx. entre clientes","Amplitude","Nº Clientes"]
            st.dataframe(tabela.style.format({
                "Faturamento":fmt_brl, "Preço Médio (CX)":fmt_brl,
                "Mín. entre clientes":fmt_brl, "Máx. entre clientes":fmt_brl,
                "Amplitude":fmt_pct, "Nº Clientes":fmt_int,
            }), use_container_width=True, hide_index=True, height=380)

        st.markdown("---")

        # ----- Drill-down por produto: ver preço de cada cliente -----
        st.subheader("🔍 Drill-down: Preço por Cliente em um Produto Específico")
        opcoes_prod = pm.sort_values("Faturamento", ascending=False)["Produto_Label"].tolist()
        prod_sel = st.selectbox("Escolha um produto para inspecionar:", opcoes_prod,
                                  help="Mostra o preço que cada cliente pagou por este produto.")
        if prod_sel:
            cod = prod_sel.split(" — ")[0]
            dist = distorcao_preco_cliente(vendas_f, cod)
            if dist.empty:
                st.info("Sem dados para esse produto.")
            else:
                preco_med = dist["Preco_CX"].median()
                col_d1, col_d2, col_d3 = st.columns(3)
                col_d1.metric("Preço Mediano", fmt_brl(preco_med))
                col_d2.metric("Cliente Mais Caro",
                                fmt_brl(dist["Preco_CX"].max()),
                                delta=fmt_delta_pct((dist["Preco_CX"].max()-preco_med)/preco_med))
                col_d3.metric("Cliente Mais Barato",
                                fmt_brl(dist["Preco_CX"].min()),
                                delta=fmt_delta_pct((dist["Preco_CX"].min()-preco_med)/preco_med))

                fig_d = px.bar(
                    dist.head(20), x="Preco_CX", y="Cliente_Label", orientation="h",
                    color="% vs Mediana", color_continuous_scale="RdBu_r",
                    color_continuous_midpoint=0,
                    text=dist.head(20)["Preco_CX"].apply(fmt_brl),
                    hover_data={"Estado": True, "Regiao_Desc": True, "Quantidade": ":,.0f"},
                )
                fig_d.add_vline(x=preco_med, line_dash="dash", line_color="#102A43",
                                 annotation_text="Mediana", annotation_position="top")
                fig_d.update_traces(textposition="outside")
                fig_d.update_layout(height=520, yaxis=dict(autorange="reversed", title=""),
                                     xaxis_title="Preço médio (R$/CX) pago pelo cliente",
                                     margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_d, use_container_width=True)

                with st.expander("📋 Tabela completa de preços por cliente"):
                    dist_view = dist[["Cliente_Label","Regiao_Desc","Estado",
                                        "Quantidade","Faturamento","Preco_CX","% vs Mediana"]]
                    dist_view.columns = ["Cliente","Região","UF","Qtd (CX)",
                                          "Faturamento","Preço (R$/CX)","% vs Mediana"]
                    st.dataframe(dist_view.style.format({
                        "Qtd (CX)":fmt_int, "Faturamento":fmt_brl,
                        "Preço (R$/CX)":fmt_brl, "% vs Mediana":fmt_delta_pct,
                    }), use_container_width=True, hide_index=True)


# -----------------------------------------------------------------------------
# TAB 6 — COMPARAÇÃO ENTRE CLIENTES/ESTADOS
# -----------------------------------------------------------------------------
with tab_comp:
    st.markdown(
        "Selecione **2 ou mais** clientes ou estados para comparar lado a lado. "
        "Útil para comparar concorrentes (ex: Sendas vs BBA) ou regiões geográficas."
    )

    tipo_comp = st.radio("O que comparar?", ["Clientes", "Estados (UF)"],
                          horizontal=True)
    dim_col = "Cliente_Label" if tipo_comp == "Clientes" else "Estado"
    opcoes = sorted([x for x in vendas_f[dim_col].dropna().unique() if str(x) != "nan"])

    # Pré-seleção dos top 3 da dimensão
    if len(opcoes) >= 2:
        tops = (vendas_f.groupby(dim_col)["Vlr.Total"].sum()
                 .sort_values(ascending=False).head(3).index.tolist())
        default_sel = [t for t in tops if t in opcoes][:3]
    else:
        default_sel = []

    entidades = st.multiselect(
        f"Escolha 2 a 6 {tipo_comp.lower()}:",
        opcoes, default=default_sel, max_selections=6)

    if len(entidades) < 2:
        st.info(f"Selecione pelo menos 2 {tipo_comp.lower()} acima para iniciar a comparação.")
    else:
        # Tabela comparativa
        st.subheader("Resumo Lado a Lado")
        comp_df = comparar_entidades(vendas_f, devol_f, dim_col, entidades)
        if comp_df is None or comp_df.empty:
            st.warning("Sem dados.")
        else:
            comp_t = comp_df.set_index("Entidade").T.reset_index()
            comp_t.columns = ["Indicador"] + list(comp_df["Entidade"])
            # Formatação por linha
            def fmt_linha(row):
                ind = row["Indicador"]
                for col in row.index[1:]:
                    val = row[col]
                    if ind in ("Faturamento Bruto","Devoluções","Líquido",
                                "Preço Médio (R$/CX)"):
                        row[col] = fmt_brl(val)
                    elif ind == "Preço Médio (R$/kg)":
                        row[col] = fmt_rkg(val)
                    elif ind in ("Qtd","Clientes Ativos","SKUs"):
                        row[col] = fmt_int(val)
                    elif ind == "% Devol.":
                        row[col] = fmt_pct(val)
                return row
            comp_show = comp_t.apply(fmt_linha, axis=1)
            st.dataframe(comp_show, use_container_width=True, hide_index=True, height=400)

            # Gráfico: evolução mensal das entidades
            st.subheader("Evolução Mensal — Faturamento Líquido")
            ev_data = []
            for ent in entidades:
                v = vendas_f[vendas_f[dim_col] == ent].groupby("AnoMes")["Vlr.Total"].sum()
                d = devol_f[devol_f[dim_col] == ent].groupby("AnoMes")["Vlr.Total"].sum()
                ml = pd.DataFrame({"V":v, "D":d}).fillna(0)
                ml["L"] = ml["V"] - ml["D"]
                ml = ml.reset_index()
                ml["Entidade"] = ent
                ml["Label"] = ml["AnoMes"].dt.strftime("%b/%y")
                ev_data.append(ml)
            ev = pd.concat(ev_data, ignore_index=True).sort_values("AnoMes")
            fig_ev = px.line(ev, x="Label", y="L", color="Entidade", markers=True,
                              color_discrete_sequence=PALETA_COMPARE)
            fig_ev.update_traces(line=dict(width=2.5), marker=dict(size=8))
            fig_ev.update_layout(height=380, yaxis_title="R$ Líquido", xaxis_title="",
                                  hovermode="x unified", margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_ev, use_container_width=True)

            # Mix de Linha lado a lado
            st.subheader("Mix por Linha de Atuação (% do Faturamento)")
            mix_data = []
            for ent in entidades:
                v_ent = vendas_f[vendas_f[dim_col] == ent]
                mix = v_ent.groupby("Linha")["Vlr.Total"].sum()
                if mix.sum() > 0:
                    mix = (mix / mix.sum() * 100).reset_index()
                    mix["Entidade"] = ent
                    mix_data.append(mix)
            if mix_data:
                mix_df = pd.concat(mix_data, ignore_index=True)
                fig_mix = px.bar(mix_df, x="Entidade", y="Vlr.Total", color="Linha",
                                   barmode="stack", color_discrete_map=COLORS,
                                   text=mix_df["Vlr.Total"].apply(lambda v: f"{v:.0f}%"))
                fig_mix.update_layout(height=380, yaxis_title="% do Faturamento",
                                        xaxis_title="", margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_mix, use_container_width=True)

            # Comparação de R$/kg
            st.subheader("Preço Médio R$/kg por Entidade")
            rkg_data = []
            for ent in entidades:
                v_ent = vendas_f[(vendas_f[dim_col] == ent) &
                                  vendas_f["Peso_kg_pack"].notna()]
                if v_ent["Peso_kg_total"].sum() > 0:
                    rkg = v_ent["Vlr.Total"].sum() / v_ent["Peso_kg_total"].sum()
                    rkg_data.append({"Entidade": ent, "R$/kg": rkg})
            if rkg_data:
                rkg_df = pd.DataFrame(rkg_data)
                fig_rk = px.bar(rkg_df, x="Entidade", y="R$/kg",
                                  color="Entidade", color_discrete_sequence=PALETA_COMPARE,
                                  text=rkg_df["R$/kg"].apply(fmt_rkg))
                fig_rk.update_traces(textposition="outside")
                fig_rk.update_layout(height=350, showlegend=False, yaxis_title="R$ por kg",
                                       xaxis_title="", margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_rk, use_container_width=True)


# -----------------------------------------------------------------------------
# TAB 7 — ANÁLISES ESTRATÉGICAS (ABC + Sazonalidade + Crescimento Geo)
# -----------------------------------------------------------------------------
with tab_estrat:
    if vendas_f.empty:
        st.warning("Sem dados.")
    else:
        # ===== ABC =====
        st.subheader("🎯 Curva ABC de Produtos — Princípio de Pareto")
        st.caption("Classe A: até 80% do faturamento · B: próximos 15% · C: últimos 5%")
        abc, total = analise_abc(vendas_f, devol_f)
        if abc.empty or total <= 0:
            st.info("Sem produtos para análise ABC.")
        else:
            cresumo = abc.groupby("Classe").agg(
                Qtd=("Produto","count"), Líquido=("Líquido","sum")).reset_index()
            cresumo["%"] = cresumo["Líquido"] / total

            cA, cB, cC = st.columns(3)
            for col, classe, emoji in [(cA,"A","🟢"), (cB,"B","🟡"), (cC,"C","🔴")]:
                row = cresumo[cresumo["Classe"] == classe]
                if not row.empty:
                    r = row.iloc[0]
                    col.metric(f"{emoji} Classe {classe}", f"{int(r['Qtd'])} produtos",
                                delta=f"{fmt_pct(r['%'])} do faturamento")

            fig_abc = go.Figure()
            chart = abc.head(25).copy()
            fig_abc.add_bar(x=chart["Produto_Label"].str[:30], y=chart["Líquido"],
                              marker_color=[COLORS[c] for c in chart["Classe"]],
                              name="Líquido")
            fig_abc.add_trace(go.Scatter(
                x=chart["Produto_Label"].str[:30],
                y=chart["% Acumulado"] * chart["Líquido"].max(),
                yaxis="y2", mode="lines+markers",
                line=dict(color="#102A43", width=2), name="% Acumulado"))
            fig_abc.update_layout(
                height=450, hovermode="x unified",
                xaxis=dict(tickangle=-45),
                yaxis=dict(title="R$ Líquido"),
                yaxis2=dict(title="% Acumulado", overlaying="y", side="right",
                             tickformat=",.0%", range=[0, chart["Líquido"].max()*1.05]),
                margin=dict(l=10,r=10,t=20,b=120))
            st.plotly_chart(fig_abc, use_container_width=True)

            with st.expander("📋 Ver tabela completa da Curva ABC"):
                tbl = abc[["Produto_Label","Vlr.Total_v","Vlr.Total_d","Líquido","%","% Acumulado","Classe"]]
                tbl.columns = ["Produto","Bruto","Devoluções","Líquido","%","% Acum.","Classe"]
                st.dataframe(tbl.style.format({
                    "Bruto":fmt_brl,"Devoluções":fmt_brl,"Líquido":fmt_brl,
                    "%":fmt_pct,"% Acum.":fmt_pct}),
                              use_container_width=True, hide_index=True)

        st.markdown("---")
        # ===== SAZONALIDADE =====
        st.subheader("📅 Sazonalidade — Heatmap Ano × Mês")
        psaz, media_mes = sazonalidade(vendas_f, devol_f)
        if not psaz.empty:
            pd_disp = psaz.copy()
            pd_disp.columns = [MESES_PT[m] for m in pd_disp.columns]
            fig_hm = px.imshow(pd_disp.values, x=pd_disp.columns.tolist(),
                                 y=[str(i) for i in pd_disp.index],
                                 color_continuous_scale="Blues", aspect="auto",
                                 labels=dict(x="Mês", y="Ano", color="R$ Líquido"),
                                 text_auto=".2s")
            fig_hm.update_layout(height=280, margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_hm, use_container_width=True)

            cs1, cs2 = st.columns([2, 1])
            with cs1:
                st.subheader("Média por Mês-do-Ano")
                fig_med = px.bar(media_mes, x="MesNome", y="Media_Liquido",
                                  color="Media_Liquido", color_continuous_scale="Blues",
                                  text=media_mes["Media_Liquido"].apply(fmt_brl))
                fig_med.update_traces(textposition="outside")
                fig_med.update_layout(height=330, showlegend=False, coloraxis_showscale=False,
                                        yaxis_title="R$ médio", xaxis_title="",
                                        margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_med, use_container_width=True)
            with cs2:
                st.subheader("Insights")
                if media_mes["Media_Liquido"].notna().any():
                    pico = media_mes.loc[media_mes["Media_Liquido"].idxmax()]
                    baixo = media_mes.loc[media_mes["Media_Liquido"].idxmin()]
                    st.success(f"📈 **Mais forte:** {pico['MesNome']}\n\n{fmt_brl(pico['Media_Liquido'])}")
                    st.warning(f"📉 **Mais fraco:** {baixo['MesNome']}\n\n{fmt_brl(baixo['Media_Liquido'])}")
                    st.info(f"🔄 **Amplitude:** {(pico['Media_Liquido']/baixo['Media_Liquido']):.2f}x")

        st.markdown("---")
        # ===== CRESCIMENTO GEO =====
        st.subheader("🌎 Crescimento Geográfico (YoY 2025 vs 2026)")
        if not {2025, 2026}.issubset(set(vendas_f["Ano"].unique())):
            st.info("YoY geográfico precisa de dados de 2025 e 2026.")
        else:
            tr, tu = st.tabs(["Por Região", "Por Estado (UF)"])
            with tr:
                gr = crescimento_geo(vendas_f, devol_f, "Regiao_Desc")
                if not gr.empty:
                    fig_gr = px.bar(gr, x="Regiao_Desc", y="Δ Abs",
                                      color="Δ Abs", color_continuous_scale="RdYlGn",
                                      text=gr["Δ Abs"].apply(fmt_brl))
                    fig_gr.update_traces(textposition="outside")
                    fig_gr.update_layout(height=350, showlegend=False,
                                            coloraxis_showscale=False,
                                            yaxis_title="Δ Líquido (R$)", xaxis_title="",
                                            margin=dict(l=10,r=10,t=20,b=10))
                    st.plotly_chart(fig_gr, use_container_width=True)
                    gr.columns = [str(c) for c in gr.columns]
                    st.dataframe(gr.style.format({
                        "2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct,
                    }), use_container_width=True, hide_index=True)
            with tu:
                gu = crescimento_geo(vendas_f, devol_f, "Estado")
                if not gu.empty:
                    gu.columns = [str(c) for c in gu.columns]
                    st.dataframe(gu.style.format({
                        "2025":fmt_brl,"2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct,
                    }), use_container_width=True, hide_index=True)


# -----------------------------------------------------------------------------
# TAB 8 — RISCOS & OPERAÇÕES
# -----------------------------------------------------------------------------
with tab_risco:
    if vendas_f.empty:
        st.warning("Sem dados.")
    else:
        # Clientes em risco
        st.subheader("⚠️ Clientes em Risco")
        st.caption("Queda ≥30% em 2026 vs mesmos meses de 2025. Considera clientes com ≥ R$ 1.000 em 2025.")
        risco = clientes_em_risco(vendas_f)
        if risco.empty:
            st.success("✅ Sem clientes em risco identificados.")
        else:
            r1, r2, r3 = st.columns(3)
            n_crit = (risco["Nível"]=="🔴 Crítico (parou)").sum()
            n_sev = (risco["Nível"]=="🔴 Severo").sum()
            n_alto = (risco["Nível"]=="🟠 Alto").sum()
            r1.metric("🔴 Críticos (pararam)", n_crit)
            r2.metric("🔴 Severos (>70%)", n_sev)
            r3.metric("🟠 Alto (>50%)", n_alto)

            r_show = risco.copy()
            if 2025 in r_show.columns: r_show = r_show.rename(columns={2025:"Fat 2025"})
            if 2026 in r_show.columns: r_show = r_show.rename(columns={2026:"Fat 2026"})
            cols = [c for c in ["Cliente_Label","Regiao_Desc","Estado",
                                  "Fat 2025","Fat 2026","Δ Abs","Δ %","Nível"] if c in r_show.columns]
            r_show = r_show[cols].rename(columns={"Cliente_Label":"Cliente"})
            st.dataframe(r_show.head(50).style.format({
                "Fat 2025":fmt_brl,"Fat 2026":fmt_brl,"Δ Abs":fmt_brl,"Δ %":fmt_delta_pct,
            }), use_container_width=True, hide_index=True, height=380)

        st.markdown("---")
        # Devoluções
        st.subheader("↩️ Devoluções Problemáticas")
        st.caption("Itens e clientes com maior taxa (mín. R$ 5k vendido).")
        pp, pc = devolucoes_problematicas(vendas_f, devol_f)
        tab_pp, tab_pc = st.tabs(["Produtos", "Clientes"])
        with tab_pp:
            if pp.empty: st.info("Sem dados suficientes.")
            else:
                top_p = pp.head(15)
                fig_pp = px.bar(top_p, x="Taxa", y=top_p["Produto_Label"].str[:40],
                                  orientation="h", color="Taxa", color_continuous_scale="Reds",
                                  text=top_p["Taxa"].apply(fmt_pct))
                fig_pp.update_traces(textposition="outside")
                fig_pp.update_layout(height=420, yaxis=dict(autorange="reversed", title=""),
                                       xaxis_title="Taxa Devolução", xaxis_tickformat=".0%",
                                       showlegend=False, coloraxis_showscale=False,
                                       margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_pp, use_container_width=True)
                st.dataframe(pp.head(30)[["Produto_Label","Vendido","Devolvido","Taxa"]].rename(
                    columns={"Produto_Label":"Produto"}).style.format({
                        "Vendido":fmt_brl,"Devolvido":fmt_brl,"Taxa":fmt_pct}),
                              use_container_width=True, hide_index=True)
        with tab_pc:
            if pc.empty: st.info("Sem dados suficientes.")
            else:
                top_c = pc.head(15)
                fig_pc = px.bar(top_c, x="Taxa", y=top_c["Cliente_Label"].str[:40],
                                  orientation="h", color="Taxa", color_continuous_scale="Reds",
                                  text=top_c["Taxa"].apply(fmt_pct))
                fig_pc.update_traces(textposition="outside")
                fig_pc.update_layout(height=420, yaxis=dict(autorange="reversed", title=""),
                                       xaxis_title="Taxa Devolução", xaxis_tickformat=".0%",
                                       showlegend=False, coloraxis_showscale=False,
                                       margin=dict(l=10,r=10,t=20,b=10))
                st.plotly_chart(fig_pc, use_container_width=True)
                st.dataframe(pc.head(30)[["Cliente_Label","Vendido","Devolvido","Taxa"]].rename(
                    columns={"Cliente_Label":"Cliente"}).style.format({
                        "Vendido":fmt_brl,"Devolvido":fmt_brl,"Taxa":fmt_pct}),
                              use_container_width=True, hide_index=True)

        st.markdown("---")
        # Vendedores
        st.subheader("👔 Ranking de Vendedores")
        vend = ranking_vendedores(vendas_f)
        if vend.empty or vend["Faturamento"].sum() == 0:
            st.info("Sem dados de vendedores.")
        else:
            v_top = vend.head(15)
            fig_v = px.bar(v_top, x="Faturamento", y="Vendedor", orientation="h",
                             color="Faturamento", color_continuous_scale="Blues",
                             text=v_top["Faturamento"].apply(fmt_brl))
            fig_v.update_traces(textposition="outside")
            fig_v.update_layout(height=430, yaxis=dict(autorange="reversed"),
                                  xaxis_title="R$ Faturamento", showlegend=False,
                                  coloraxis_showscale=False, margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_v, use_container_width=True)
            st.dataframe(vend.style.format({
                "Faturamento":fmt_brl,"Quantidade":fmt_int,
                "Pedidos":fmt_int,"Ticket_Medio":fmt_brl,
            }), use_container_width=True, hide_index=True)


# -----------------------------------------------------------------------------
# TAB 9 — DADOS DETALHADOS
# -----------------------------------------------------------------------------
with tab_dados:
    st.subheader("Vendas (após filtros)")
    cols_v = [c for c in ["Emissao","Cliente_Label","Regiao_Desc","Estado","Municipio",
                            "Produto_Label","Linha","Categoria","Materia_Prima",
                            "Vendedor","Quantidade","Peso_kg_total","Vlr.Total"] if c in vendas_f.columns]
    show_v = vendas_f[cols_v].rename(columns={"Cliente_Label":"Cliente","Produto_Label":"Produto"})
    st.dataframe(show_v.sort_values("Emissao", ascending=False),
                  use_container_width=True, hide_index=True, height=320)
    st.download_button("⬇️ Vendas (CSV)", show_v.to_csv(index=False).encode("utf-8"),
                       file_name="vendas_filtradas.csv", mime="text/csv")

    st.markdown("---")
    st.subheader("Devoluções (após filtros)")
    cols_d = [c for c in ["Emissao","Cliente_Label","Regiao_Desc","Estado","Municipio",
                            "Produto_Label","Linha","Categoria","Materia_Prima",
                            "Quantidade","Vlr.Total"] if c in devol_f.columns]
    show_d = devol_f[cols_d].rename(columns={
        "Emissao":"DT Digitacao","Cliente_Label":"Cliente","Produto_Label":"Produto"})
    st.dataframe(show_d.sort_values("DT Digitacao", ascending=False),
                  use_container_width=True, hide_index=True, height=320)
    st.download_button("⬇️ Devoluções (CSV)", show_d.to_csv(index=False).encode("utf-8"),
                       file_name="devolucoes_filtradas.csv", mime="text/csv")

st.markdown("---")
st.caption(
    "💡 Devoluções pelo mês da DT Digitacao (regra contábil). "
    "Peso/kg extraído da descrição. Filtros em cascata. "
    "Cliente e Produto exibidos como `Código — Nome`."
)
